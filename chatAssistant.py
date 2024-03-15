import base64

import openai
import streamlit as st
from bs4 import BeautifulSoup
from faker.decode import unidecode
from streamlit.components.v1 import html
import requests
import pdfkit
import time
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from PIL import Image
import urllib
import pandas as pd

load_dotenv()
#id do assistente
assistant_id = "asst_xtYbsxpLumPBvxrytbr1fex4"

# inicializa cliente openai
client = openai

# inicializa a sessão para ler os ids
if "file_id_list" not in st.session_state:
    st.session_state.file_id_list = []

if "start_chat" not in st.session_state:
    st.session_state.start_chat = False

if "thread_id" not in st.session_state:
    st.session_state.thread_id = None

# titulo e icone da página
# Função para converter XLSX pra PDF

def convert_xlsx_to_json(input_path, output_path) :
    read_file = pd.read_excel(input_path)

    read_file.to_json(output_path)

def convert_xlsx_to_pdf(input_path, output_path):
    workbook = load_workbook(input_path)
    sheets = workbook.sheetnames

    pdf = canvas.Canvas(output_path, pagesize=letter)

    for sheet_name in sheets:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                pdf.drawString(cell.column * 50, letter[1] - cell.row * 10, str(cell.value))

    pdf.save()


perguntas = [
    "Qual é a distribuição de pedidos ao longo dos anos e meses?",
    "Quais são os produtos mais pedidos e em quais categorias eles se encaixam?",
    "Existe alguma diferença significativa na quantidade de pedidos entre os diferentes dias da semana?",
    "Qual a distribuição de gênero e faixa etária dos clientes que fizeram pedidos?",
    "Quais são as cidades e unidades federativas com mais pedidos?",
    "Qual é o valor médio dos pedidos aprovados em comparação com os pedidos não aprovados?",
    "Qual é a distribuição das marcas mais populares nos pedidos?",
    "Existe alguma sazonalidade nas compras de acordo com o mês ou semana do ano?",
    "É possível identificar alguma tendência ou padrão nos valores dos pedidos?",
    "Existe alguma correlação entre a quantidade de itens pedidos e o valor dos pedidos aprovados?"
]

pergunta_ = ""

st.sidebar.write('<style>p, ol, ul, dl {font-size:0.9rem}</style>', unsafe_allow_html=True)

def getImage(file_id) :
    image_file_id = file_id
    image_file = openai.files.content(image_file_id)
    bites = BytesIO(base64.b64decode(image_file.content))
    aux_im = Image.open(BytesIO(image_file.content))
    return aux_im

def download_file(file) :
    file = urllib.request.urlopen(file).read()
    return BytesIO(file)

# Função pra enviar arquivo convertido pra OpenAI
def upload_to_openai(filepath):
    with open(filepath, "rb") as file:
        response = openai.files.create(file=file.read(), purpose="assistants")
    return response.id

#local
#api_key = os.getenv("OPENAI_API_KEY")
#git
api_key = st.secrets.OpenAIAPI.openai_api_key

if api_key:
    openai.api_key = api_key

#st.sidebar.write("<a style='color:white'  href='https://tecnologia2.chleba.net/_ftp/chatgpt/BotasVentoPedidos.xlsx' id='baixarArquivo'>[Baixe o arquivo para fazer a análise]</a>", unsafe_allow_html=True)

#uploaded_file = st.sidebar.file_uploader("Envie um arquivo", key="file_uploader")
#uploaded_file = download_file("https://tecnologia2.chleba.net/_ftp/chatgpt/BotasVentoPedidos.xlsx")
# Botão para iniciar o chat

if not st.session_state.start_chat:
    if True:
        #if uploaded_file:
        ds = client.beta.assistants.files.list(assistant_id=assistant_id)
        if ds:
            for file in ds:
                st.session_state.file_id_list.append(file.id)
            #st.sidebar.write(f"ID do arquivo: {additional_file_id}")

        # Mostra os ids
        if st.session_state.file_id_list:
            #st.sidebar.write("IDs dos arquivos enviados:")
            for file_id in st.session_state.file_id_list:
                #st.sidebar.write(file_id)
                # Associa os arquivos ao assistente
                assistant_file = client.beta.assistants.files.create(
                    assistant_id=assistant_id,
                    file_id=file_id
                )


        # Verifica se o arquivo foi upado antes de iniciar
        if st.session_state.file_id_list:
            st.session_state.start_chat = True
            # Cria a thread e guarda o id na sessão
            thread = client.beta.threads.create()
            st.session_state.thread_id = thread.id
            #st.write("id da thread: ", thread.id)
        else:
            st.sidebar.warning("Por favor, clique em \"Iniciar análise\" iniciar o chat")


if st.session_state.start_chat:
    on = st.sidebar.toggle('Ver sugestões de perguntas', value=True)
    search = st.sidebar.text_input("Pesquisar perguntas sugeridas")

    if on:
        for indice, pergunta in enumerate(perguntas):
            # st.sidebar.write(f"<a style=\"color:white;display:flex;align-items:center;gap:26px;text-decoration:none\" target=\"_self\" id=\"pergunta{indice}\" href=\"javascript:(function(){{var conteudo = document.getElementById('pergunta{indice}').innerText; navigator.clipboard.writeText(conteudo).then(function() {{ console.log('Conteúdo copiado para a área de transferência: ' + conteudo); }}, function(err) {{ console.error('Erro ao copiar conteúdo: ', err); }});}})()\">{pergunta}<span>{icon_copy}</span></a>", unsafe_allow_html=True)
            if unidecode(search.lower()) in unidecode(pergunta.lower()):
                if st.sidebar.button(f"{pergunta}"):
                    pergunta_ = pergunta

    st.sidebar.write('<style>label[data-baseweb="checkbox"] > div > div {background: #282828}</style>', unsafe_allow_html=True)
# Define a função para iniciar
def process_message_with_citations(message):
    """Extract content and annotations from the message and format citations as footnotes."""
    message_content = message.content[0].text
    annotations = message_content.annotations if hasattr(message_content, 'annotations') else []
    citations = []

    # for nas annotations
    for index, annotation in enumerate(annotations):
        # substitui o texto da mensagem
        message_content.value = message_content.value.replace(annotation.text, f' [{index + 1}]')

        if (file_citation := getattr(annotation, 'file_citation', None)):
            cited_file = {'filename': 'cited_document.pdf'}  # Substituído pelo arquivo retornado
            citations.append(f'[{index + 1}] {file_citation.quote} from {cited_file["filename"]}')
        elif (file_path := getattr(annotation, 'file_path', None)):
            # Placeholder for file download citation
            cited_file = {'filename': 'downloaded_document.pdf'}  # Substituído pelo arquivo retornado
            citations.append(f'[{index + 1}] Click [here](#) to download {cited_file["filename"]}')  # Link de download substituído pelo caminho do arquivo

    # Adiciona notas no final da mensgaem (talvez tirar)
    full_response = message_content.value
    return full_response

# Interface do chat
st.subheader("ANÁLISE DE PRODUTOS")
#st.write("Este chat usa a API da OpenAI para gerar respostas.")

# Só vai mostrar o chat se for iniciado
if st.session_state.start_chat:
    # Inicializa o modelo usado
    if "openai_model" not in st.session_state:
        st.session_state.openai_model = "gpt-4-turbo-preview"
    if "messages" not in st.session_state:
        st.session_state.messages = []

    # Mostra mensagens anteriores
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            if message["typeFile"] == "text":
                st.markdown(message["content"], unsafe_allow_html=True)

            if message["typeFile"] == "image":
                image = getImage(message["content"])
                st.image(image)


    prompt_ =  st.chat_input("Faça uma pergunta!" )

    if pergunta_ :
        prompt = pergunta_

    if not pergunta_ :
        prompt = prompt_

    # Campo pro usuário escrever
    if prompt:
        # Adiciona as mensagens do usuário e mostra no chat
        st.session_state.messages.append({"role": "user", "content": prompt, "typeFile":"text"})
        with st.chat_message("user"):
            st.markdown(prompt, unsafe_allow_html=True)

        # Adiciona as mensagens criadas na thread
        client.beta.threads.messages.create(
            thread_id=st.session_state.thread_id,
            role="user",
            content=prompt,
            file_ids=st.session_state.file_id_list
        )

        # Cria a requisição com mais instruções
        run = client.beta.threads.runs.create(
            thread_id=st.session_state.thread_id,
            assistant_id=assistant_id,
            instructions="Por favor, responda as perguntas usando o conteúdo do arquivo. Quando adicionar informações externas, seja claro e mostre essas informações em outra cor."
        )


        # Pedido para finalizar a requisição e retornar as mensagens do assistente
        while run.status != 'completed':
            time.sleep(1)
            run = client.beta.threads.runs.retrieve(
                thread_id=st.session_state.thread_id,
                run_id=run.id
            )



        # Retorna as mensagens do assistente
        messages = client.beta.threads.messages.list(
            thread_id=st.session_state.thread_id
        )

        # Processa e mostra as mensagens do assistente
        assistant_messages_for_run = [
            message for message in messages
            if message.run_id == run.id and message.role == "assistant"
        ]
        for message in assistant_messages_for_run[::-1]:
            #print(message)
            if message.content[0].type == "text":
                full_response = process_message_with_citations(message)
                st.session_state.messages.append({"role": "assistant", "content": full_response, "typeFile" :"text"})

                with st.chat_message("assistant"):
                    st.markdown(full_response, unsafe_allow_html=True)

            if message.content[0].type == "image_file":
                image = getImage(message.content[0].image_file.file_id)
                if image:
                    st.session_state.messages.append({"role": "assistant", "content": message.content[0].image_file.file_id, "typeFile" : "image"})
                    with st.chat_message("assistant"):
                        st.image(getImage(message.content[0].image_file.file_id))




else:
    # Prompt pra iniciar o chat
    st.write("Por favor, selecione o(s) arquivo(s) e clique em *iniciar chat* para gerar respostas")
